"""
Voucher Filtration Automation — Shopee
=======================================
Run: streamlit run app.py

PID-level rule: if ANY SKU under a Product ID has a REAL eligibility failure
(Status=NO, future launch date, AM excluded, or price below threshold),
ALL SKUs under that Product ID are blanked.

Key Shopee logic vs Lazada:
  - SKUs with Status=#N/A (not in zeCOM tracking) are treated as
    "untracked / no restriction" — they can still get YES for mechanics
    that include '#N/A' in their match values (e.g. 50% NMS open default).
  - Untracked SKUs do NOT cause PID blocking.
"""

import streamlit as st
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import io
import warnings
from datetime import date, datetime
from collections import defaultdict

warnings.filterwarnings("ignore")

NA = "#N/A"
TODAY = date.today()

st.set_page_config(page_title="Shopee Voucher Filtration", page_icon="🟠", layout="wide")

st.markdown("""
<style>
    .block-container { padding-top: 2rem; }
    .stButton > button {
        background-color: #ee4d2d; color: white; font-weight: 700;
        border: none; padding: 0.6rem 2rem; border-radius: 6px;
        font-size: 16px; width: 100%;
    }
    .stButton > button:hover { background-color: #ff6347; }
</style>
""", unsafe_allow_html=True)

st.title("🟠 Shopee Voucher Filtration")
st.caption(f"PID-level campaign filtering  ·  Run date: {TODAY}")
st.divider()

tab_run, tab_clear = st.tabs(["▶  Run Filtration", "🗑️  Clear Output Columns"])

# ── SIDEBAR ────────────────────────────────────────────────────
with st.sidebar:
    st.header("⚙️ Settings")

    st.subheader("Eligibility Rules")
    with st.expander("How PID-level blocking works", expanded=False):
        st.markdown("""
**Pass 1 — per-SKU classification:**
- `Status = NO` → **blocks PID**
- `Status = YES` + future launch date → **blocks PID**
- `Status = YES` + price below threshold → **blocks PID**
- `Status = YES` + AM excluded → **blocks PID**
- `Status = YES` + passes all → eligible, matched by Exclusion text
- `Status = #N/A` (not in zeCOM) → **does NOT block PID**, eligible for mechanics that include `#N/A` in match values

**Pass 2 — PID gate:**
If ANY SKU in a Product ID triggered a block in Pass 1,
ALL SKUs in that Product ID get blank mechanic columns.
        """)

    apply_threshold = st.checkbox("Require RRP & SRP ≥ minimum price", value=True)
    min_price = st.number_input("Minimum price", value=39, min_value=0, step=1,
        disabled=not apply_threshold)

    st.divider()
    st.subheader("Campaign Type")
    st.caption("Auto-selects the correct zeCOM price columns")
    CAMPAIGN_PRICE_COLS = {
        "BAU":               {"price": 54, "special_price": 55, "disc_pct": 56},
        "Mid Month / Payday":{"price": 57, "special_price": 58, "disc_pct": 59},
        "Double Digit":      {"price": 60, "special_price": 61, "disc_pct": 62},
        "Mega":              {"price": 63, "special_price": 64, "disc_pct": 65},
    }
    campaign_type = st.selectbox("Shopee campaign type",
        list(CAMPAIGN_PRICE_COLS.keys()), index=1)
    price_cols = CAMPAIGN_PRICE_COLS[campaign_type]

    st.divider()
    st.subheader("zeCOM Column Positions")
    st.caption("0-based index")
    col_style  = st.number_input("Style# (ALU_NO)",  value=2,  min_value=0)
    col_launch = st.number_input("Launch Date",       value=21, min_value=0)
    col_status = st.number_input("Status Shopee",     value=24, min_value=0)
    col_excl   = st.number_input("Exclusion",         value=66, min_value=0)
    st.caption(f"Price cols auto-set ({campaign_type}): "
               f"RRP={price_cols['price']} SRP={price_cols['special_price']} Disc={price_cols['disc_pct']}")

    st.divider()
    st.subheader("Shopee File Column Names")
    st.caption("Must match your SellerPriceTemplate row 1 headers exactly")
    sku_col_name   = st.text_input("SKU / EAN column",      value="SKU")
    price_col_name = st.text_input("Price column",           value="Price")
    pid_col_name   = st.text_input("Product ID column",      value="Product ID")
    data_start_row = st.number_input("Data starts at row",   value=2, min_value=2)

ZECOM_COLS = {
    "style": int(col_style), "launch_date": int(col_launch),
    "status": int(col_status), "exclusion": int(col_excl),
    "price": price_cols["price"], "special_price": price_cols["special_price"],
    "disc_pct": price_cols["disc_pct"],
}


# ── HELPERS ────────────────────────────────────────────────────
def normalise_ean(v):
    return str(v).strip().split(".")[0].strip()


def matches_mechanic(excl_val, mech):
    """Returns True if excl_val matches this mechanic's rules."""
    el = str(excl_val).strip().lower()
    mv = [v.lower() for v in mech["match_values"]]
    ex = [v.lower() for v in mech.get("excludes", [])]
    matched = any(m in el for m in mv) if mech["match_type"] == "contains" else el in mv
    return matched and not any(x in el for x in ex if x)


def build_mechanics(raw):
    out = []
    for m in raw:
        name = m["name"].strip()
        mvs = [v.strip() for v in m["match_values"].splitlines() if v.strip()]
        if name and mvs:
            out.append({
                "name": name,
                "match_type": m["match_type"],
                "match_values": mvs,
                "excludes": [v.strip() for v in m["excludes"].split(",") if v.strip()],
            })
    return out


def classify_sku(status, live, alu_no, rrp, srp, am_set):
    """
    Returns one of:
      'eligible'    — Status=YES, passes all checks
      'status_no'   — Status=NO  → blocks PID
      'future'      — future launch date → blocks PID
      'am_excluded' — in AM exclusion list → blocks PID
      'low_price'   — RRP or SRP below threshold → blocks PID
      'untracked'   — Status=#N/A (not in zeCOM) → does NOT block PID,
                      eligible only for mechanics with '#N/A' in match values
    """
    if status == "NO":
        return "status_no"
    if status == NA or status is None:
        return "untracked"
    # Status = YES from here
    if isinstance(live, (date, datetime)):
        ld = live.date() if isinstance(live, datetime) else live
        if ld > TODAY:
            return "future"
    if alu_no and alu_no in am_set:
        return "am_excluded"
    if apply_threshold:
        for pv in (rrp, srp):
            if pv not in (NA, None):
                try:
                    if float(pv) < min_price:
                        return "low_price"
                except (TypeError, ValueError):
                    pass
    return "eligible"


def mechanic_ui(state_key):
    if state_key not in st.session_state:
        st.session_state[state_key] = [
            {"name": "", "match_type": "contains", "match_values": "", "excludes": ""}
        ]
    _remove = None
    for i, mech in enumerate(st.session_state[state_key]):
        with st.container(border=True):
            c1, c2, c3 = st.columns([4, 2, 1])
            mech["name"] = c1.text_input(
                "Column header name", value=mech["name"],
                key=f"{state_key}_name_{i}",
                placeholder="e.g. 50% NMS (50% VC ONLY - Shopee exclusive clearance)")
            mech["match_type"] = c2.selectbox(
                "Match type", ["contains", "exact"],
                index=["contains", "exact"].index(mech["match_type"]),
                key=f"{state_key}_type_{i}")
            c3.markdown("&nbsp;")
            if c3.button("🗑️", key=f"{state_key}_rm_{i}"):
                _remove = i
            mech["match_values"] = st.text_area(
                "zeCOM Exclusion value(s) — one per line  "
                "*(add `#N/A` to include untracked/not-in-zeCOM products)*",
                value=mech["match_values"], key=f"{state_key}_mv_{i}", height=80,
                placeholder="50% VC ONLY - Shopee exclusive clearance\n#N/A")
            mech["excludes"] = st.text_input(
                "Exclude if Exclusion text also contains (comma-separated, optional)",
                value=mech["excludes"], key=f"{state_key}_ex_{i}",
                placeholder="e.g. No platform VC")
    if _remove is not None:
        st.session_state[state_key].pop(_remove)
        st.rerun()
    if st.button("➕ Add mechanic", key=f"{state_key}_add"):
        st.session_state[state_key].append(
            {"name": "", "match_type": "contains", "match_values": "", "excludes": ""})
        st.rerun()


# ── TAB 1: RUN FILTRATION ──────────────────────────────────────
with tab_run:
    st.subheader("🎯 Voucher Mechanics")
    st.info(
        "**Tip for open/default mechanics (e.g. 50% NMS):** add `#N/A` as one of the "
        "match values to include products not in the zeCOM tracking file. "
        "These products have no specific exclusion restriction and default to the open campaign.",
        icon="💡"
    )
    mechanic_ui("sh_mechanics")

    st.divider()
    st.subheader("📂 Upload Input Files")
    c1, c2, c3 = st.columns(3)
    with c1:
        st.markdown("**1. Price / Stock File** *(required)*")
        price_file = st.file_uploader("Shopee SellerPriceTemplate",
            type=["xlsx", "xls"], key="price")
    with c2:
        st.markdown("**2. Content File** *(required)*")
        content_file = st.file_uploader("Content_file",
            type=["xlsx", "xls"], key="content")
    with c3:
        st.markdown("**3. zeCOM Tracking File** *(required)*")
        zecom_file = st.file_uploader("zeCOM_Tracking_File",
            type=["xlsx", "xls"], key="zecom")
    c4, c5 = st.columns(2)
    with c4:
        st.markdown("**4. AM Exclusion File** *(optional)*")
        am_file = st.file_uploader("AM_Exclusion.xlsx",
            type=["xlsx", "xls"], key="am")
    with c5:
        st.markdown("**Output filename** *(optional)*")
        out_name = st.text_input("out", value="", label_visibility="collapsed",
            placeholder="e.g. Shopee_Payday_June_filtration.xlsx", key="out_name")

    st.divider()
    mechanics = build_mechanics(st.session_state.get("sh_mechanics", []))
    ready = price_file and content_file and zecom_file and mechanics
    if not ready:
        st.info("👆 Upload 3 required files and define at least one mechanic to get started")

    if st.button("▶  Run Shopee Filtration", disabled=not ready):
        with st.spinner("Processing (PID-level, 2-pass)..."):
            try:
                logs = []
                def log(m): logs.append(m)

                # ── Content map ───────────────────────────────
                content_file.seek(0)
                cwb = openpyxl.load_workbook(content_file, read_only=True, data_only=True)
                if "content" not in cwb.sheetnames:
                    raise ValueError("Content file must have a sheet named 'content'.")
                cws = cwb["content"]
                ch = [c.value for c in next(cws.iter_rows(min_row=1, max_row=1))]
                ean_i   = [str(h).strip().lower() for h in ch].index("ean")
                color_i = [str(h).strip().lower() for h in ch].index("color_no")
                content_map = {}
                for row in cws.iter_rows(min_row=2, values_only=True):
                    if row[ean_i] is not None:
                        content_map[normalise_ean(row[ean_i])] = str(row[color_i]).strip()
                cwb.close()
                log(f"✓ {len(content_map):,} EAN → ALU_NO mappings")

                # ── zeCOM map ─────────────────────────────────
                zecom_file.seek(0)
                zwb = openpyxl.load_workbook(zecom_file, read_only=True, data_only=True)
                if "MY" not in zwb.sheetnames:
                    raise ValueError("zeCOM file must have a sheet named 'MY'.")
                zws = zwb["MY"]
                zecom_map = {}
                for row in zws.iter_rows(min_row=5, values_only=True):
                    style = row[ZECOM_COLS["style"]] if len(row) > ZECOM_COLS["style"] else None
                    if not style: continue
                    def g(k, row=row):
                        i = ZECOM_COLS.get(k)
                        return row[i] if i is not None and len(row) > i else None
                    zecom_map[str(style).strip()] = {
                        "Launch_Date":  g("launch_date"), "Status":       g("status"),
                        "Price":        g("price"),       "Special_Price":g("special_price"),
                        "Disc_Pct":     g("disc_pct"),    "Exclusion":    g("exclusion"),
                    }
                zwb.close()
                log(f"✓ {len(zecom_map):,} styles from zeCOM ({campaign_type})")

                # ── AM exclusion ──────────────────────────────
                am_set = set()
                if am_file:
                    am_file.seek(0)
                    awb = openpyxl.load_workbook(am_file, read_only=True, data_only=True)
                    for row in awb.active.iter_rows(min_row=2, values_only=True):
                        if row and row[0]: am_set.add(str(row[0]).strip())
                    awb.close()
                log(f"✓ {len(am_set):,} AM-excluded ALU_NOs")

                # ── Load Price/Stock workbook ─────────────────
                price_file.seek(0)
                wb = openpyxl.load_workbook(price_file)
                ws = wb[wb.sheetnames[0]]
                headers = [c.value for c in ws[1]]

                def fc(name):
                    for idx, h in enumerate(headers, 1):
                        if h and str(h).strip().lower() == name.lower(): return idx
                    return None

                sku_col   = fc(sku_col_name)
                price_col = fc(price_col_name)
                pid_col   = fc(pid_col_name)
                if not sku_col:
                    raise ValueError(f"Column '{sku_col_name}' not found in row 1. "
                                     f"Available: {[h for h in headers if h]}")
                if not price_col:
                    raise ValueError(f"Column '{price_col_name}' not found in row 1.")
                if not pid_col:
                    raise ValueError(f"Column '{pid_col_name}' not found in row 1. "
                                     f"Check 'Product ID column' setting in sidebar.")
                last_col = max((i for i,h in enumerate(headers,1) if h), default=len(headers))
                log(f"✓ Sheet='{ws.title}' | SKU=col{sku_col} PID=col{pid_col} Price=col{price_col}")

                # ── PASS 1: classify each SKU ─────────────────
                # classification: 'eligible' | 'untracked' | 'status_no' |
                #                 'future' | 'am_excluded' | 'low_price' | 'no_content'
                pass1 = []
                total = matched_alu = matched_zecom = 0
                N_LOOKUP = 8

                for r in range(int(data_start_row), ws.max_row + 1):
                    sku = ws.cell(row=r, column=sku_col).value
                    if sku is None: continue
                    total += 1
                    orig_price = ws.cell(row=r, column=price_col).value
                    pid = ws.cell(row=r, column=pid_col).value
                    alu_no = content_map.get(normalise_ean(sku))

                    if alu_no is None:
                        row_vals = [NA] * N_LOOKUP
                        clf = "no_content"
                    else:
                        matched_alu += 1
                        rec = zecom_map.get(alu_no)
                        if rec is None:
                            row_vals = [alu_no] + [NA] * (N_LOOKUP - 1)
                            clf = "untracked"
                        else:
                            matched_zecom += 1
                            live   = rec["Launch_Date"]
                            status = rec["Status"]
                            rrp    = rec["Price"]
                            srp    = rec["Special_Price"]
                            pct    = rec["Disc_Pct"]
                            excl   = rec["Exclusion"]
                            try:
                                orig_p = float(str(orig_price).replace("'","").strip()) \
                                         if orig_price is not None else None
                                rrp_check = round(float(rrp),2) == round(orig_p,2) \
                                            if orig_p is not None else NA
                            except: rrp_check = NA
                            row_vals = [alu_no, live, status, rrp, srp, rrp_check, pct, excl]
                            clf = classify_sku(status, live, alu_no, rrp, srp, am_set)
                            # also treat Status=#N/A from zeCOM record as untracked
                            if status in (NA, None) or status == NA:
                                clf = "untracked"

                    excl_v = row_vals[7]
                    pass1.append((r, pid, alu_no, row_vals, excl_v, clf))

                log(f"✓ Pass 1: {total:,} SKUs | {matched_alu:,} ALU_NO | {matched_zecom:,} zeCOM")

                # ── PASS 2: PID gate ──────────────────────────
                # Only real failures block a PID; 'untracked' and 'no_content' do not
                BLOCKING = {"status_no", "future", "am_excluded", "low_price"}
                pid_elig = defaultdict(lambda: True)
                for _, pid, _, _, _, clf in pass1:
                    if clf in BLOCKING:
                        pid_elig[pid] = False

                pid_total   = len(pid_elig)
                pid_ok      = sum(1 for v in pid_elig.values() if v)
                pid_blocked = pid_total - pid_ok
                log(f"✓ Pass 2: {pid_total:,} PIDs | {pid_ok:,} eligible | {pid_blocked:,} blocked")

                # ── Write output ──────────────────────────────
                new_headers = (
                    ["ALU_NO","Live","Status","RRP","SRP","RRP check","%","Exclusions"]
                    + [m["name"] for m in mechanics]
                )
                base_font   = Font(name="Aptos Narrow", size=11, bold=False)
                yellow_fill = PatternFill(fill_type="solid", fgColor="FFFFFF00")

                for i, h in enumerate(new_headers):
                    col  = last_col + 1 + i
                    cell = ws.cell(row=1, column=col, value=h)
                    cell.font = base_font
                    if i >= N_LOOKUP: cell.fill = yellow_fill
                    ws.column_dimensions[get_column_letter(col)].width = max(
                        ws.column_dimensions[get_column_letter(col)].width or 0, 16)

                counts = {m["name"]: 0 for m in mechanics}

                for r, pid, alu_no, row_vals, excl_v, clf in pass1:
                    pid_ok_flag = pid_elig.get(pid, True)
                    mech_vals = []
                    for m in mechanics:
                        if clf == "no_content":
                            # No ALU match → blank everything
                            mech_vals.append("")
                        elif not pid_ok_flag:
                            # PID blocked → blank
                            mech_vals.append("")
                        elif clf in BLOCKING:
                            # Individually ineligible → blank
                            mech_vals.append("")
                        elif clf == "untracked":
                            # Not in zeCOM → match only if '#N/A' is in match_values
                            yes = matches_mechanic(NA, m)
                            if yes: counts[m["name"]] += 1
                            mech_vals.append("YES" if yes else "")
                        else:
                            # clf == 'eligible' → normal Exclusion text matching
                            yes = matches_mechanic(excl_v, m)
                            if yes: counts[m["name"]] += 1
                            mech_vals.append("YES" if yes else "")

                    for i, val in enumerate(row_vals + mech_vals):
                        col  = last_col + 1 + i
                        cell = ws.cell(row=r, column=col, value=val)
                        cell.font = base_font
                        if i == 1 and isinstance(val, (date, datetime)):
                            cell.number_format = "mm-dd-yy"

                for n, c in counts.items(): log(f"  {n}: {c:,} YES")

                buf = io.BytesIO()
                wb.save(buf)
                buf.seek(0)

                # ── Results ───────────────────────────────────
                st.success("✅ Shopee filtration complete!")
                mc = st.columns(4 + len(mechanics))
                with mc[0]: st.metric("Total SKUs",     f"{total:,}")
                with mc[1]: st.metric("ALU_NO matched", f"{matched_alu:,}")
                with mc[2]: st.metric("PIDs eligible",  f"{pid_ok:,}")
                with mc[3]: st.metric("PIDs blocked",   f"{pid_blocked:,}")
                for i,(n,c) in enumerate(counts.items()):
                    with mc[4+i]: st.metric(n[:22], f"{c:,}")

                with st.expander("📝 Processing Log"):
                    for l in logs: st.text(l)

                fname = (out_name.strip()
                         or price_file.name.rsplit(".",1)[0] + "_shopee_filtration.xlsx")
                st.download_button("⬇️  Download Shopee Output", buf, fname,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True)

            except Exception as e:
                st.error(f"❌ {e}"); st.exception(e)


# ── TAB 2: CLEAR OUTPUT COLUMNS ────────────────────────────────
with tab_clear:
    st.subheader("🗑️ Clear Output Columns from Processed File")
    st.caption("Strips ALU_NO and all appended columns, returning the file to the "
               "original Shopee SellerPriceTemplate.")
    st.divider()

    clear_file = st.file_uploader("Upload processed Excel to clear",
        type=["xlsx","xls"], key="clear_upload")

    if clear_file:
        clear_file.seek(0)
        try:
            wb_p = openpyxl.load_workbook(clear_file, read_only=True)
            ws_p = wb_p[wb_p.sheetnames[0]]
            row1 = [ws_p.cell(row=1, column=c).value
                    for c in range(1, ws_p.max_column + 1)]
            wb_p.close()

            strip_from = next(
                (i+1 for i,h in enumerate(row1)
                 if h and str(h).strip().upper() == "ALU_NO"), None)

            if strip_from:
                n_strip  = len(row1) - strip_from + 1
                stripped = [str(row1[c-1]) for c in range(strip_from, len(row1)+1)]
                st.success(
                    f"✅ Found ALU_NO at column **{get_column_letter(strip_from)}** "
                    f"— will remove **{n_strip}** column(s).")
                with st.expander("Columns to be removed"):
                    for name in stripped: st.markdown(f"- `{name}`")
                st.divider()
                if st.button("🗑️  Strip & prepare download", use_container_width=True):
                    with st.spinner("Removing columns..."):
                        clear_file.seek(0)
                        wb_out = openpyxl.load_workbook(clear_file)
                        wb_out[wb_out.sheetnames[0]].delete_cols(strip_from, n_strip)
                        buf = io.BytesIO(); wb_out.save(buf); buf.seek(0)
                    clean_name = (clear_file.name
                                  .replace("_shopee_filtration","")
                                  .rsplit(".",1)[0] + "_clean.xlsx")
                    st.download_button("⬇️  Download Clean File", buf, clean_name,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True)
                    st.success(f"✅ Done — {n_strip} column(s) removed.")
            else:
                st.warning("⚠️ No ALU_NO column found — file may already be clean.")
        except Exception as e:
            st.error(f"Could not open file: {e}")
