"""
Voucher Filtration Automation — Shopee CLI
==========================================
Campaign rules and file paths live in config.py.

PID-level rule: if ANY SKU under a Product ID is ineligible,
ALL SKUs under that Product ID are blanked.

Output: the Shopee SellerPriceTemplate with ALU_NO / Live / Status /
RRP / SRP / RRP check / % / Exclusions + one YES/blank column per mechanic
appended directly onto the same sheet.

Usage:
    python voucher_filtration.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import os
import sys
import warnings
from datetime import date, datetime
from collections import defaultdict

warnings.filterwarnings('ignore')

try:
    from config import (
        PRICE_FILE, CONTENT_FILE, ZECOM_FILE, AM_EXCL_FILE, OUTPUT_FILE,
        PRODUCT_ID_COL_NAME, DATA_START_ROW,
        ZECOM_COLS, MIN_PRICE_THRESHOLD, APPLY_MIN_PRICE_THRESHOLD, CAMPAIGNS,
        CAMPAIGN_TYPE
    )
except ImportError as e:
    print(f"ERROR: Could not load config.py — {e}")
    sys.exit(1)

NA = '#N/A'
TODAY = date.today()

print("=" * 60)
print("  SHOPEE VOUCHER FILTRATION")
print(f"  Campaign type : {CAMPAIGN_TYPE}")
print(f"  Run date      : {TODAY}")
print("=" * 60)


def normalise_ean(v):
    return str(v).strip().split('.')[0].strip()


def matches_campaign(excl_val, camp):
    el = str(excl_val).strip().lower()
    mv = [v.lower() for v in camp['match_values']]
    ex = [v.lower() for v in camp.get('excludes', [])]
    matched = any(m in el for m in mv) if camp['match_type'] == 'contains' else el in mv
    return matched and not any(x in el for x in ex if x)


def is_eligible(status, live, alu_no, rrp, srp, am_set):
    if status != 'YES':
        return False
    if isinstance(live, (date, datetime)):
        ld = live.date() if isinstance(live, datetime) else live
        if ld > TODAY:
            return False
    if alu_no and alu_no in am_set:
        return False
    if APPLY_MIN_PRICE_THRESHOLD:
        for pv in (rrp, srp):
            if pv not in (NA, None):
                try:
                    if float(pv) < MIN_PRICE_THRESHOLD:
                        return False
                except (TypeError, ValueError):
                    pass
    return True


# ============================================================
# STEP 1: Load Price/Stock File
# ============================================================
print('\n▶ STEP 1: Loading Price/Stock file...')
if not os.path.exists(PRICE_FILE):
    print(f'  ERROR: File not found → {PRICE_FILE}'); sys.exit(1)

wb = openpyxl.load_workbook(PRICE_FILE)
ws = wb[wb.sheetnames[0]]
headers = [c.value for c in ws[1]]
print(f"  ✓ Loaded '{ws.title}' sheet ({ws.max_row:,} rows)")

def find_col(name):
    for idx, h in enumerate(headers, 1):
        if h and str(h).strip().lower() == name.lower():
            return idx
    return None

sku_col   = find_col('SellerSKU')
price_col = find_col('Price')
pid_col   = find_col(PRODUCT_ID_COL_NAME)

if not sku_col:  print('  ERROR: No SellerSKU column in row 1'); sys.exit(1)
if not price_col: print('  ERROR: No Price column in row 1'); sys.exit(1)
if not pid_col:
    print(f"  ERROR: No '{PRODUCT_ID_COL_NAME}' column in row 1 — update PRODUCT_ID_COL_NAME in config.py")
    sys.exit(1)

print(f"  ✓ SellerSKU=col{sku_col} | Price=col{price_col} | ProductID('{PRODUCT_ID_COL_NAME}')=col{pid_col}")
last_col = max((i for i,h in enumerate(headers,1) if h), default=len(headers))


# ============================================================
# STEP 2: Load Content File
# ============================================================
print('\n▶ STEP 2: Mapping ALU_NO via Content file...')
if not os.path.exists(CONTENT_FILE):
    print(f'  ERROR: File not found → {CONTENT_FILE}'); sys.exit(1)

cwb = openpyxl.load_workbook(CONTENT_FILE, read_only=True, data_only=True)
cws = cwb['content']
ch = [c.value for c in next(cws.iter_rows(min_row=1, max_row=1))]
ean_i   = [str(h).strip().lower() for h in ch].index('ean')
color_i = [str(h).strip().lower() for h in ch].index('color_no')
content_map = {}
for row in cws.iter_rows(min_row=2, values_only=True):
    if row[ean_i] is not None:
        content_map[normalise_ean(row[ean_i])] = str(row[color_i]).strip()
cwb.close()
print(f'  ✓ {len(content_map):,} EAN → ALU_NO mappings')


# ============================================================
# STEP 3: Load zeCOM Tracking File
# ============================================================
print(f'\n▶ STEP 3: Loading zeCOM Tracking file ({CAMPAIGN_TYPE})...')
if not os.path.exists(ZECOM_FILE):
    print(f'  ERROR: File not found → {ZECOM_FILE}'); sys.exit(1)

zwb = openpyxl.load_workbook(ZECOM_FILE, read_only=True, data_only=True)
zws = zwb['MY']
zecom_map = {}
for row in zws.iter_rows(min_row=5, values_only=True):
    style = row[ZECOM_COLS['style']] if len(row) > ZECOM_COLS['style'] else None
    if not style: continue
    def g(k, row=row):
        i = ZECOM_COLS.get(k)
        return row[i] if i is not None and len(row) > i else None
    zecom_map[str(style).strip()] = {
        'Launch_Date': g('launch_date'), 'Status': g('status'),
        'Price': g('price'), 'Special_Price': g('special_price'),
        'Disc_Pct': g('disc_pct'), 'Exclusion': g('exclusion'),
    }
zwb.close()
print(f'  ✓ {len(zecom_map):,} styles loaded')


# ============================================================
# STEP 4: AM Exclusion (optional)
# ============================================================
am_set = set()
if AM_EXCL_FILE and os.path.exists(AM_EXCL_FILE):
    print('\n▶ STEP 4: Loading AM Exclusion file...')
    awb = openpyxl.load_workbook(AM_EXCL_FILE, read_only=True, data_only=True)
    for row in awb.active.iter_rows(min_row=2, values_only=True):
        if row and row[0]: am_set.add(str(row[0]).strip())
    awb.close()
    print(f'  ✓ {len(am_set):,} excluded ALU_NOs')
else:
    print('\n▶ STEP 4: No AM Exclusion file — skipped')


# ============================================================
# STEP 5: PASS 1 — per-SKU eligibility
# ============================================================
print('\n▶ STEP 5: Pass 1 — checking per-SKU eligibility...')

pass1 = []  # (r, pid, alu_no, row_vals, excl_v, elig)
total = matched_alu = matched_zecom = 0
N_LOOKUP = 8

for r in range(DATA_START_ROW, ws.max_row + 1):
    sku = ws.cell(row=r, column=sku_col).value
    if sku is None: continue
    total += 1
    orig_price = ws.cell(row=r, column=price_col).value
    pid = ws.cell(row=r, column=pid_col).value
    alu_no = content_map.get(normalise_ean(sku))

    if alu_no is None:
        row_vals = [NA] * N_LOOKUP
    else:
        matched_alu += 1
        rec = zecom_map.get(alu_no)
        if rec is None:
            row_vals = [alu_no] + [NA] * (N_LOOKUP - 1)
        else:
            matched_zecom += 1
            live, status = rec['Launch_Date'], rec['Status']
            rrp, srp, pct, excl = rec['Price'], rec['Special_Price'], rec['Disc_Pct'], rec['Exclusion']
            try: rrp_check = round(float(rrp), 2) == round(float(orig_price), 2)
            except: rrp_check = NA
            row_vals = [alu_no, live, status, rrp, srp, rrp_check, pct, excl]

    status_v, live_v, rrp_v, srp_v, excl_v = row_vals[2], row_vals[1], row_vals[3], row_vals[4], row_vals[7]
    elig = (excl_v != NA) and is_eligible(status_v, live_v, alu_no, rrp_v, srp_v, am_set)
    pass1.append((r, pid, alu_no, row_vals, excl_v, elig))

print(f'  ✓ {total:,} SKUs | {matched_alu:,} ALU_NO matched | {matched_zecom:,} zeCOM matched')


# ============================================================
# STEP 6: PASS 2 — PID-level gate
# ============================================================
print('\n▶ STEP 6: Pass 2 — applying PID-level rule...')

pid_elig = defaultdict(lambda: True)
for _, pid, _, _, _, elig in pass1:
    if not elig:
        pid_elig[pid] = False

pid_total   = len(pid_elig)
pid_ok      = sum(1 for v in pid_elig.values() if v)
pid_blocked = pid_total - pid_ok
print(f'  ✓ {pid_total:,} Product IDs | {pid_ok:,} fully eligible | {pid_blocked:,} blocked by PID rule')


# ============================================================
# STEP 7: Write output columns
# ============================================================
print('\n▶ STEP 7: Writing output columns...')

new_headers = ['ALU_NO','Live','Status','RRP','SRP','RRP check','%','Exclusions'] \
    + [c['name'] for c in CAMPAIGNS]
base_font   = Font(name='Aptos Narrow', size=11, bold=False)
yellow_fill = PatternFill(fill_type='solid', fgColor='FFFFFF00')

for i, h in enumerate(new_headers):
    col  = last_col + 1 + i
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = base_font
    if i >= N_LOOKUP: cell.fill = yellow_fill
    ws.column_dimensions[get_column_letter(col)].width = max(
        ws.column_dimensions[get_column_letter(col)].width or 0, 14)

counts = {c['name']: 0 for c in CAMPAIGNS}

for r, pid, alu_no, row_vals, excl_v, sku_elig in pass1:
    final_elig = sku_elig and pid_elig.get(pid, True)
    mech_vals = []
    for camp in CAMPAIGNS:
        if not final_elig:
            mech_vals.append('')
        else:
            yes = matches_campaign(excl_v, camp)
            if yes: counts[camp['name']] += 1
            mech_vals.append('YES' if yes else '')

    for i, val in enumerate(row_vals + mech_vals):
        col  = last_col + 1 + i
        cell = ws.cell(row=r, column=col, value=val)
        cell.font = base_font
        if i == 1 and isinstance(val, (date, datetime)):
            cell.number_format = 'mm-dd-yy'

for name, count in counts.items():
    print(f'  {name}: {count:,} YES')


# ============================================================
# STEP 8: Save
# ============================================================
print(f'\n▶ STEP 8: Saving → {OUTPUT_FILE}')
os.makedirs(os.path.dirname(OUTPUT_FILE) if os.path.dirname(OUTPUT_FILE) else '.', exist_ok=True)
wb.save(OUTPUT_FILE)

print(f'\n{"="*60}')
print(f'  ✅ DONE!  Output → {OUTPUT_FILE}')
print(f'  Total SKUs     : {total:,}')
print(f'  PIDs eligible  : {pid_ok:,}')
print(f'  PIDs blocked   : {pid_blocked:,}')
for name, count in counts.items():
    print(f'  {name}: {count:,}')
print(f'  Run date       : {TODAY}')
print(f'{"="*60}\n')
